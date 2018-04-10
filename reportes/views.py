from orders.models import Order
from django.db import connection
from django.shortcuts import render, redirect,render_to_response
from .forms import RangoFechaForm
from django.core.urlresolvers import reverse
from django.http import HttpResponse,HttpResponseRedirect
from django.views.generic.base import TemplateView
from openpyxl import Workbook
import json
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.template import RequestContext
from django.contrib import messages



def loginRep(request):
	if not request.user.is_anonymous():
		return HttpResponseRedirect('/reportes/consumo')
	if request.method == 'POST':
		form = AuthenticationForm(request.POST)
		if form.is_valid:
			user= request.POST['username']
			passw = request.POST['password']
			access = authenticate(username=user,password=passw)
			request.session.set_expiry(500)
			if access is not None:
				login(request,access)

				form = RangoFechaForm()
				return HttpResponseRedirect('/reportes/consumo')
			else:
				messages.add_message(request, messages.ERROR, "*Usuario o contraseña incorrectos")
				return HttpResponseRedirect(reverse( 'reportes:login' ))
	else:
		form = AuthenticationForm()
	return render_to_response('Login.html', {'form': form}, context_instance = RequestContext(request))


@login_required(login_url='/reportes')
def Reportes(request):
	form = RangoFechaForm()

	if request.method == 'POST':
		form = RangoFechaForm(request.POST)
		if form.is_valid():
			data = form.cleaned_data
			desde = data['Desde']
			hasta = data['Hasta']
			request.session['Desde']= str(desde)
			request.session['Hasta']= str(hasta)
			cursor = connection.cursor()
			query = cursor.execute('select nombre, numerotarjeta, numeroempleado,nomina, razon_social,area,segmento,sucursal,correo, SUM(totalcompra) As total from orders_order where created >"%s" AND created < "%s" group by nombre' % (desde , hasta))
			reportes = cursor.fetchall()
			return render(request, 'reporte.html', {'reportes': reportes, 'form':form})
	else:
		return render(request, 'reporte.html', {'form':form})
	return desde,hasta

@login_required(login_url='/reportes')
def reportecomedor(request):
	form = RangoFechaForm()

	if request.method == 'POST':
		form = RangoFechaForm(request.POST)
		if form.is_valid():
			data = form.cleaned_data
			desde = data['Desde']
			hasta = data['Hasta']
			request.session['Desde']= str(desde)
			request.session['Hasta']= str(hasta)
			cursor = connection.cursor()
			query = cursor.execute('select a.name, SUM(b.quantity) AS VentaTotal FROM orders_orderitem B JOIN comedor_product A ON b.product_id = a.id JOIN orders_order c ON c.id = b.order_id WHERE c.created > "%s" AND c.created < "%s" GROUP by product_id' % (desde , hasta))
			reportes = cursor.fetchall()
			return render(request, 'reporteComedor.html', {'reportes': reportes, 'form':form})
	else:
		return render(request, 'reporteComedor.html', {'form':form})

@login_required(login_url='/reportes')
def buscaempleado(request):
		form = RangoFechaForm()
		empleado = request.GET['search']
		request.session['empleado']= str(empleado)
		cursor = connection.cursor()
		query = cursor.execute('select c.nombre, a.name, b.quantity, b.price, DATE(c.created) AS Fecha FROM orders_orderitem B JOIN comedor_product A ON b.product_id = a.id JOIN orders_order c ON c.id =  b.order_id WHERE c.numeroempleado = "%s"' % empleado)
		if query > 0:	
			reportes = cursor.fetchall()
			nombre = reportes[0][0]
			return render(request, 'reporteempleado.html', {'reportes': reportes, 'nombre': nombre, 'form':form})
		else:
			return render(request,'busqueda.html', {'form':form})	

@login_required(login_url='/reportes')
def ReporteEmpleado(request):
	form = RangoFechaForm()

	if request.method == 'POST':
		form = RangoFechaForm(request.POST)
		if form.is_valid():
			data = form.cleaned_data
			desde = data['Desde']
			hasta = data['Hasta']
			request.session['Desde']= str(desde)
			request.session['Hasta']= str(hasta)
			empleado= request.session['empleado']
			cursor = connection.cursor()
			query = cursor.execute('select c.nombre, a.name, b.quantity, b.price, DATE(c.created) AS Fecha FROM orders_orderitem B JOIN comedor_product A ON b.product_id = a.id JOIN orders_order c ON c.id =  b.order_id WHERE c.numeroempleado = "%s" AND c.created > "%s" AND c.created < "%s"' % (empleado,desde,hasta))
			reportes = cursor.fetchall()
			return render(request, 'ReporteEmpleado.html', {'reportes': reportes, 'form':form})
	else:
		return render(request, 'ReporteEmpleado.html', {'form':form})




@login_required(login_url='/reportes')
def ReporteTotalExcel(request):
		desde= request.session['Desde']
		hasta= request.session['Hasta']
		cursor = connection.cursor()
		query = cursor.execute('select nombre, numerotarjeta, numeroempleado,nomina, razon_social, SUM(totalcompra) As total from orders_order where created >"%s" AND created < "%s" group by nombre' % (desde , hasta))
		reportes = cursor.fetchall()
		wb = Workbook()
		ws = wb.active
		ws['B1'] = 'REPORTE DE GASTO TOTAL POR EMPLEADO'
		ws.merge_cells('B1:E1')
		ws['B3'] = 'NOMBRE'
		ws['C3'] = 'NUMERO TARJETA'
		ws['D3'] = 'NUMERO EMPLEADO'
		ws['E3'] = 'NOMINA'
		ws['F3'] = 'RAZON SOCIAL'  
		ws['G3'] = 'TOTAL DE COMPRA'         
		cont=4

		for reporte in reportes:
			ws.cell(row=cont,column=2).value = reporte[0]
			ws.cell(row=cont,column=3).value = reporte[1]
			ws.cell(row=cont,column=4).value = reporte[2]
			ws.cell(row=cont,column=5).value = reporte[3]
			ws.cell(row=cont,column=6).value = reporte[4]
			ws.cell(row=cont,column=7).value = reporte[5]
			
			cont = cont + 1

			#Establecemos el nombre del archivo
		nombre_archivo ="ReporteTotalEmpleado.xlsx"
		#Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
		response = HttpResponse(content_type="application/ms-excel") 
		contenido = "attachment; filename={0}".format(nombre_archivo)
		response["Content-Disposition"] = contenido
		wb.save(response)
		return response

@login_required(login_url='/reportes')
def ReporteTotalComedor(request):
		cursor = connection.cursor()
		query = cursor.execute('select a.name, SUM(b.quantity) AS VentaTotal FROM orders_orderitem B JOIN comedor_product A ON b.product_id = a.id JOIN orders_order c ON c.id = b.order_id GROUP by product_id' )
		reportes = cursor.fetchall()
		wb = Workbook()
		ws = wb.active
		ws['B1'] = 'REPORTE DE GASTO TOTAL POR EMPLEADO'
		ws.merge_cells('B1:E1')
		ws['B3'] = 'PRODUCTO'
		ws['C3'] = 'VENTA TOTAL'    
		cont=4

		for reporte in reportes:
			ws.cell(row=cont,column=2).value = reporte[0]
			ws.cell(row=cont,column=3).value = reporte[1]
			
			cont = cont + 1

			#Establecemos el nombre del archivo
		nombre_archivo ="ReporteTotalComedor.xlsx"
		#Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
		response = HttpResponse(content_type="application/ms-excel") 
		contenido = "attachment; filename={0}".format(nombre_archivo)
		response["Content-Disposition"] = contenido
		wb.save(response)
		return response


@login_required(login_url='/reportes')
def cerrar(request):
	logout(request)
	return HttpResponseRedirect('/reportes/')